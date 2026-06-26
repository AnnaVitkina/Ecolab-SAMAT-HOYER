"""Convert a cleaned rate DataFrame into the target matrix-style XLSX layout."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from project_paths import OUTPUT_DIR

BOLD_SHIPMENT_HEADERS = {
    "Carrier Name",
    "Origin Country",
    "Origin Postal Code",
    "Destination Country",
    "Destination postal code",
}

HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
BOLD = Font(bold=True)
NORMAL = Font()
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

CARRIER_NAME_MAP = {
    "Hoyer": "HOYER GMBH NSAP",
    "Hoyer GmbH": "HOYER GMBH NSAP",
    "Hoyer España SA": "HOYER ESPANA S.A. NSAP",
    "Samat Haanpa - Tesjoki": "SAMAT INTERNATIONAL NSAP",
    "Samat Normandie": "SAMAT NORMANDIE EBS",
    "Samat Spain": "SAMAT ESPANA NSAP",
}
DEFAULT_CURRENCY = "EUR"
DEFAULT_APPLY_IF = "Applies if: invoiced by Carrier"

COST_NAME_ROW = 1
APPLY_IF_ROW = 2
RATE_BY_ROW = 3
COLUMN_HEADER_ROW = 4
DATA_START_ROW = 5


@dataclass(frozen=True)
class ShipmentColumn:
    header: str
    source: str | None = None


@dataclass(frozen=True)
class CostBlock:
    title: str
    source_column: str
    apply_if: str = DEFAULT_APPLY_IF
    rate_by: str = "Rate by: Per shipment"


SHIPMENT_COLUMNS = [
    ShipmentColumn("Lane #"),
    ShipmentColumn("Lane ID", "Lane"),
    ShipmentColumn("Carrier Name", "_carrier_name"),
    ShipmentColumn("Vendor Name", "Vendor Name"),
    ShipmentColumn("Vendor Number", "Vendor Number"),
    ShipmentColumn("Origin Country", "Origin Country"),
    ShipmentColumn("Origin Postal Code", "Origin ZIP Code"),
    ShipmentColumn("Origin City", "Origin City"),
    ShipmentColumn("Destination Country", "Destination Country"),
    ShipmentColumn("Destination postal code", "Destination ZIP Code"),
    ShipmentColumn("Destination City", "Destination City"),
    ShipmentColumn("Mode", "Mode"),
]

EBS_SHIPMENT_COLUMNS = [
    ShipmentColumn("Lane #"),
    ShipmentColumn("Origin Country", "Origin Country"),
    ShipmentColumn("Origin Postal Code", "Origin ZIP Code"),
    ShipmentColumn("Origin City", "Origin City"),
    ShipmentColumn("Destination Country", "Destination Country"),
    ShipmentColumn("Destination postal code", "Destination ZIP Code"),
    ShipmentColumn("Destination City", "Destination City"),
]


def get_shipment_columns(
    df: pd.DataFrame,
    *,
    is_ebs: bool = False,
) -> list[ShipmentColumn]:
    template = EBS_SHIPMENT_COLUMNS if is_ebs else SHIPMENT_COLUMNS
    visible: list[ShipmentColumn] = []
    for column in template:
        if column.header == "Lane #":
            visible.append(column)
            continue
        if column.source == "_carrier_name":
            if "Vendor Name" in df.columns:
                visible.append(column)
            continue
        if column.source == "Vendor Number":
            if has_vendor_numbers(df):
                visible.append(column)
            continue
        if column.source and column.source not in df.columns:
            continue
        visible.append(column)
    return visible


COST_BLOCKS = [
    CostBlock("Transport cost", "Freight Rate"),
    CostBlock("Infrastructure Fee", "Surcharge infraestructure Hoyer"),
    CostBlock("ETS Fee", "ETS 2026 Hoyer"),
    CostBlock("Toll Fee", "2026 Toll BE-NL"),
    CostBlock("Cleaning costs", "Cleaning costs"),
]

EBS_COST_BLOCKS = [
    CostBlock("Transport cost", "Freight Rate"),
]


def get_cost_blocks(
    df: pd.DataFrame,
    *,
    is_ebs: bool = False,
) -> list[CostBlock]:
    template = EBS_COST_BLOCKS if is_ebs else COST_BLOCKS
    return [block for block in template if block.source_column in df.columns]


def map_carrier_name(vendor_name: object) -> str:
    if pd.isna(vendor_name):
        return ""
    text = str(vendor_name).strip()
    return CARRIER_NAME_MAP.get(text, text)


def _is_na_vendor_number(value: object) -> bool:
    if pd.isna(value):
        return True
    text = str(value).strip().upper()
    return text in {"", "N/A", "#N/A", "NA", "NAN"}


def has_vendor_numbers(df: pd.DataFrame) -> bool:
    if "Vendor Number" not in df.columns:
        return False
    return any(not _is_na_vendor_number(value) for value in df["Vendor Number"])


def _sanitize_sheet_name(name: str) -> str:
    text = re.sub(r"[\[\]:*?/\\]", " ", name).strip()
    return (text or "rates")[:31]


def _unique_sheet_name(name: str, used_names: set[str]) -> str:
    base = _sanitize_sheet_name(name)
    candidate = base
    suffix = 1
    while candidate in used_names:
        tail = f"_{suffix}"
        candidate = f"{base[: 31 - len(tail)]}{tail}"
        suffix += 1
    used_names.add(candidate)
    return candidate


def count_unique_vendors(df: pd.DataFrame) -> list[str]:
    if "Vendor Name" not in df.columns:
        return []

    vendor_series = df["Vendor Name"].dropna().astype(str).str.strip()
    vendor_series = vendor_series[vendor_series != ""]
    return sorted(vendor_series.unique().tolist(), key=str.casefold)


def prompt_split_by_vendor(df: pd.DataFrame) -> bool:
    vendors = count_unique_vendors(df)
    if len(vendors) <= 1:
        return False

    print(f"\nFound {len(vendors)} vendor names:")
    for vendor in vendors:
        print(f"  - {vendor}")

    while True:
        answer = input("Split layout into separate tabs per vendor? [yes/no]: ").strip().lower()
        if answer in {"", "yes", "y"}:
            return True
        if answer in {"no", "n"}:
            return False
        print("Invalid selection. Enter yes or no.")


def split_df_by_vendor(
    df: pd.DataFrame,
    *,
    split_by_vendor: bool = True,
) -> list[tuple[str, pd.DataFrame]]:
    if not split_by_vendor or "Vendor Name" not in df.columns:
        return [("rates", df.reset_index(drop=True))]

    vendors = count_unique_vendors(df)
    if len(vendors) <= 1:
        return [("rates", df.reset_index(drop=True))]

    used_names: set[str] = set()
    groups: list[tuple[str, pd.DataFrame]] = []
    for vendor in vendors:
        subset = df[df["Vendor Name"].astype(str).str.strip() == vendor].copy()
        subset = subset.reset_index(drop=True)
        sheet_name = _unique_sheet_name(map_carrier_name(vendor) or vendor, used_names)
        groups.append((sheet_name, subset))
    return groups


def _has_value(value: object) -> bool:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return False
    return str(value).strip() != ""


def _format_cell_value(value: object) -> object:
    if pd.isna(value):
        return None
    if isinstance(value, float) and value == int(value):
        return int(value)
    return value


def _shipment_value(df: pd.DataFrame, row_idx: int, column: ShipmentColumn) -> object:
    if column.header == "Lane #":
        return row_idx + 1
    if column.source == "_carrier_name":
        return map_carrier_name(df.at[row_idx, "Vendor Name"])
    if column.source == "Vendor Number":
        value = df.at[row_idx, "Vendor Number"]
        return None if _is_na_vendor_number(value) else _format_cell_value(value)
    if column.source:
        if column.source not in df.columns:
            return None
        return _format_cell_value(df.at[row_idx, column.source])
    return None


def _write_shipment_headers(ws, shipment_columns: list[ShipmentColumn]) -> None:
    for col_idx, column in enumerate(shipment_columns, start=1):
        cell = ws.cell(row=COLUMN_HEADER_ROW, column=col_idx, value=column.header)
        cell.font = BOLD if column.header in BOLD_SHIPMENT_HEADERS else NORMAL
        cell.fill = HEADER_FILL
        cell.alignment = LEFT


def _write_cost_block_headers(ws, block: CostBlock, start_col: int) -> None:
    end_col = start_col + 1

    for row_idx, text in (
        (COST_NAME_ROW, block.title),
        (APPLY_IF_ROW, block.apply_if),
        (RATE_BY_ROW, block.rate_by),
    ):
        ws.merge_cells(
            start_row=row_idx,
            start_column=start_col,
            end_row=row_idx,
            end_column=end_col,
        )
        cell = ws.cell(row=row_idx, column=start_col, value=text)
        cell.font = NORMAL
        cell.fill = HEADER_FILL
        cell.alignment = LEFT

    for col_idx, label in ((start_col, "Currency"), (end_col, "Flat")):
        cell = ws.cell(row=COLUMN_HEADER_ROW, column=col_idx, value=label)
        cell.font = NORMAL
        cell.fill = HEADER_FILL
        cell.alignment = LEFT


def write_layout_sheet(
    ws: Worksheet,
    df: pd.DataFrame,
    *,
    is_ebs: bool = False,
) -> None:
    if df.empty:
        raise ValueError("Cannot build layout from an empty DataFrame.")

    shipment_columns = get_shipment_columns(df, is_ebs=is_ebs)
    cost_blocks = get_cost_blocks(df, is_ebs=is_ebs)

    if not cost_blocks:
        raise ValueError(
            "No cost columns available for layout conversion "
            f"(expected at least '{COST_BLOCKS[0].source_column}')."
        )

    shipment_count = len(shipment_columns)
    _write_shipment_headers(ws, shipment_columns)

    cost_col = shipment_count + 1
    for block in cost_blocks:
        _write_cost_block_headers(ws, block, cost_col)
        cost_col += 2

    for row_idx in range(len(df)):
        excel_row = DATA_START_ROW + row_idx
        for col_idx, column in enumerate(shipment_columns, start=1):
            ws.cell(
                row=excel_row,
                column=col_idx,
                value=_shipment_value(df, row_idx, column),
            )

        cost_col = shipment_count + 1
        for block in cost_blocks:
            amount = _format_cell_value(df.at[row_idx, block.source_column])
            ws.cell(row=excel_row, column=cost_col + 1, value=amount)
            if _has_value(amount):
                ws.cell(row=excel_row, column=cost_col, value=DEFAULT_CURRENCY)
            cost_col += 2

    for col_idx in range(1, cost_col):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18


def build_layout_workbook(
    df: pd.DataFrame,
    *,
    is_ebs: bool = False,
    split_by_vendor: bool = True,
) -> Workbook:
    if df.empty:
        raise ValueError("Cannot build layout from an empty DataFrame.")

    vendor_groups = split_df_by_vendor(df, split_by_vendor=split_by_vendor)
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    for sheet_name, vendor_df in vendor_groups:
        ws = wb.create_sheet(title=sheet_name)
        write_layout_sheet(ws, vendor_df, is_ebs=is_ebs)

    return wb


def save_layout_xlsx(
    df: pd.DataFrame,
    source_path: Path,
    *,
    is_ebs: bool = False,
    split_by_vendor: bool = True,
) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / f"{source_path.stem}_layout.xlsx"
    wb = build_layout_workbook(
        df,
        is_ebs=is_ebs,
        split_by_vendor=split_by_vendor,
    )
    wb.save(output_path)
    return output_path
